# coding=utf-8
from docx import Document
import shutil
from shutil import copyfile
import os
import subprocess
import stat
import datetime
import pyexcel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import ctypes  # An included library with Python install.

print "Modules imported."

destination_xl = 'Follow-up_CR_CN_ECU_Daimler_solenoide_piezo.xlsx'
destination_sheet_save_as = 'Follow-up_CR_CN_ECU_Daimler_solenoide_piezo.xlsx'
destination_sheet_name = 'DAIMLER 651'
code_folder_name = "CODE"

print "Constants defined."

item_quantity = 0
values_to_xl = []
mismatch_files = []

blank_row = (' ', ' ', ' ', ' ', ' ')

#for reading some datas from freigabe file and return title, ref_daimler, sw_title, sw_ref_daimler, folder name, s19 file name.
def reach_to_docx(doc_name, grouping):

	if doc_name.endswith('.docx'):
		document = Document(doc_name)
		tables = document.tables

		if not tables[0].rows[31].cells[2].text:

			if not tables[0].rows[32].cells[2].text:

				if not tables[0].rows[33].cells[2].text:

					if not tables[0].rows[34].cells[2].text:
						key_position = 0
						add_error_log(doc_name + " file has no code.")
					else:
						unique_key = tables[0].rows[34].cells[2].text
						key_position = 4
				else:
					unique_key = tables[0].rows[33].cells[2].text
					key_position = 3
			else:
				unique_key = tables[0].rows[32].cells[2].text
				key_position = 2

			if key_position != 0:
				add_error_log(doc_name + " 's code is in " + str(key_position) + ". line.")

		else:
			unique_key = tables[0].rows[31].cells[2].text
			key_position = 1

		if key_position != 0:
			unique_key_file_name = unique_key + '.s19'

			if not grouping:
				daimler_part_number = doc_name.split("/")[1].split("_")[1]
			else:
				daimler_part_number = doc_name.split("_")[1]

			file_name_first = "-".join(unique_key.split("-")[3:-1])
			file_name_sec = unique_key.split("-")[-1].split("_")[0]
			folder_name = file_name_first + '-' + file_name_sec + '-' + daimler_part_number

			sw_update_code = tables[0].rows[6].cells[3].text

			title = "-".join(unique_key.split("-")[0:3]) + '-' + daimler_part_number
			ref_daimler = daimler_part_number

			sw_ref_daimler = sw_update_code.replace(" ", "")
			sw_title = "-".join(unique_key.split("-")[0:3]) + '-' + sw_ref_daimler

	if grouping:
		return unique_key_file_name, folder_name, doc_name

	else:
		return title, ref_daimler, sw_title, sw_ref_daimler

#it adds values of (values_to_xl) array to excel and makes style settings.
def to_excel():
	destination_book = openpyxl.load_workbook(destination_xl)
	destination_sheet = destination_book[destination_sheet_name]
	print destination_xl + " found and opened."

	last_row = destination_sheet.max_row
	i = 1
	# <-----List Elements to Excel and Styling--------->

	for row in values_to_xl:
		destination_sheet.append(row)
		for col in range(1, 12):
			destination_sheet.cell(row=last_row + i, column=col).alignment = aligment_type
			destination_sheet.cell(row=last_row + i, column=col).border = my_border
			if delivery == 1:
				destination_sheet.cell(row=last_row + i, column=col).fill = delivery_color

		destination_sheet.cell(row=last_row + i, column=2).number_format = 'DD.MM.YYYY'

		destination_sheet.cell(row=last_row + i, column=21).alignment = aligment_type
		destination_sheet.cell(row=last_row + i, column=21).border = my_border
		if delivery == 1:
			destination_sheet.cell(row=last_row + i, column=21).fill = delivery_color

		i += 1

	destination_sheet.append(blank_row)

	for col in range(1, 12):
		destination_sheet.cell(row=last_row + i, column=col).fill = blank_row_color
		destination_sheet.cell(row=last_row + i, column=col).border = my_border

	destination_sheet.cell(row=last_row + i, column=21).fill = blank_row_color
	destination_sheet.cell(row=last_row + i, column=21).border = my_border

	# <----------------End of Blank Row Part----------->

	# <----------------Save File----------------------->

	destination_book.save(destination_sheet_save_as)

#for grouping process. it makes create folder and moves required s19 docx files in.
def create_folder_and_move(unique_key_file_name, folder_name, file_name): 
	if os.path.exists(unique_key_file_name):
		if not os.path.exists(folder_name):
			os.makedirs(folder_name)

			shutil.move(unique_key_file_name, folder_name)
			shutil.move(file_name, folder_name)

	else:
		add_error_log(file_name + " 's code dont match with\n " + unique_key_file_name)
		   
#message box function
def message_box(title, message):
	ctypes.windll.user32.MessageBoxA(None, message, title, 0)

#get indus number from folder name(automatic)
def get_indus_number_from_folder():
	indus_folder_name = os.getcwd().split("\\")[-1]

	if indus_folder_name.startswith("ME"):
		indus_number = indus_folder_name
	elif indus_folder_name.startswith("Indus"):
		indus_number = indus_folder_name.split("Indus")[-1]
	return int(indus_number) 

#get indus number from user (manuel)
def get_indus_number(): 
	print "--------------------------------------------------------------"
	print "PLEASE ENTER INDUS NUMBER: "
	i_num = raw_input()
	return i_num

#it checks whether the number of files in the given folder(folder) is equal to the given parameter(num)
def check_num_of_files(folder, num):
	num_of_docx = len([name for name in os.listdir(folder) if name.endswith('.docx') or name.endswith('.doc')])
	num_of_s19 = len([name for name in os.listdir(folder) if name.endswith('.s19') or name.endswith('.S19')])

	if num_of_docx == num and num_of_s19 == num:
		return True
	else:
		return False

#checks whether the indus number with sw_versinon has already been added to the excel table
def check_appended_before(indus, sw_num):
	destination_book = openpyxl.load_workbook(destination_xl)
	destination_sheet = destination_book[destination_sheet_name]

	for i in range(1, destination_sheet.max_row):
		if destination_sheet.cell(row=i, column=5).value == sw_num and destination_sheet.cell(row=i,column=6).value == indus:
			append_before = True
			break
		else:
			append_before = False

	return append_before

#return the required data for the given index number(delivery code, ecu type, sw version, cr number, cr creation)
def get_from_xl(indus):

	source_sheet = pyexcel.get_sheet(file_name=source_xl)
	source_sheet.name_columns_by_row(0)
	print source_xl + " found and opened."

	for row in source_sheet:
		if row[3] == int(indus):
			found_in_xl_ = True
			##delivery_ = row[2]
			##ecu_type_ = row[4][3:]
			##sw_version_ = row[6]
			##cr_number_ = row[8]
			##cr_creation_ = row[10]
			##return True, row[2], row[4][3:], row[6], row[8], row[10]
			break

		else:
			found_in_xl_=False
	return found_in_xl_, row[2], row[4][3:], row[6], row[8], row[10]

#setting row background color
def set_row_color(color_code):
	return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

#setting border style of row
def set_border_style(style):
	return Border(left=Side(style=style), right=Side(style=style), top=Side(style=style), bottom=Side(style=style))

#setting aligment type of row
def set_aligment (horizontal, vertical):
	return openpyxl.styles.Alignment(horizontal=horizontal,vertical=horizontal)

#for giving read-write permission to the given file
def remove_file_permission(file):
	os.chmod(file, stat.S_IWRITE)
	print "File read and write permission granted."

#it adds values to array with blank columns
def add_list(cr_num, cr_crea, sw_ver, indus, ecu, titl, ref_daim, item_quantity, comment):
	values_to_xl.append((cr_num, cr_crea, '', '', sw_ver, indus, ecu, titl, ref_daim, '', item_quantity, '', '', '', '', '', '', '', '', '', comment))

#checks whether the values ​​sent to the function are empty
def check_missing_value(*args):
	for arg in args:
		if arg == "" or arg ==" ":
			result = True
			break
		else:
		  result = False
	return result

#for creating code folder
def create_code_folder():
	src = filter(lambda x: os.path.isdir(x), os.listdir('.'))
	os.makedirs('CODE')
	in_f = os.listdir(src[0])
	for file in in_f:
		if file.endswith(".docx"):
			shutil.copy(src[0]+'/'+file, 'CODE')
	
#checks whether the sw_version in the excel table matches the file
def check_sw_version(sw_ver):
	src = filter(lambda x: os.path.isdir(x), os.listdir('.'))
	f_name = src[0]
	for files in os.listdir(f_name):
		if files.endswith('.s19') or  files.endswith('.S19'):
			s19_sw = files.split("-")[2]
			if sw_ver == s19_sw:
				return True
			else:
				return False

#for converting doc files to docx files
def convert_doc_to_docx(path):
	print "Doc files converting to docx."
	ffff = os.listdir(path)
	for fi in ffff:
		if fi.endswith('.doc'):
			subprocess.call(['C:\Program Files\LibreOffice\program\soffice.exe', '--headless', '--convert-to', 'docx', path +'/'+ fi, '--outdir',path])
	print "Converting process completed."		
#Ems table name can vary according to years so this function detect name of ems table.
def find_ems_table_name():
	for file in os.listdir(os.curdir):
		if file.startswith("EMS_Meeting_Minutes_"):
			return file
			break

#for adding error_log records to log file and save them as a txt file 
def add_error_log(error_message):
	with open("Error-Logs.txt", "a+") as text_file:
				text_file.write(error_message + '\n')
				text_file.write("---------------------------------------------------------------------------\n")

def remove_old_doc_files(path):
	for file in os.listdir(path):
		if file.endswith('.doc'):
			os.remove(path+'/'+file)

source_xl = find_ems_table_name()

my_border = set_border_style('thin') 
aligment_type = set_aligment('center', 'center') 

blank_row_color = set_row_color('a6a6a6') 
delivery_color = set_row_color('fde9d9') 

convert_doc_to_docx(os.curdir)
remove_old_doc_files(os.curdir)

folder_list = filter(lambda x: os.path.isdir(x), os.listdir('.')) #list of folders in main folder
for folders in folder_list:
	convert_doc_to_docx(folders)
	remove_old_doc_files(folders)

is_require_grouping = not check_num_of_files(os.curdir, 0) 
is_exist_source_xl = os.path.isfile(source_xl)
is_exist_destination_xl = os.path.isfile(destination_xl) 

if is_require_grouping: #if files require grouping, this statement makes grouping.
	print "Files are grouping" 
	for file_list in os.listdir(os.curdir):
		if file_list.endswith('.docx'):
			s19, fold, docx = reach_to_docx(file_list, True)
			if os.path.isfile(s19):
				create_folder_and_move(s19, fold, docx)
			else:
				mismatch_files.append(s19)
			
	print "Files grouping process finished."

is_ready_for_to_xl = check_num_of_files(os.curdir, 0) 

if not is_ready_for_to_xl: #if there is ungrouped files, error messages will appear and will not proceed to later stages of process.
	print "There are non-grouped files. Please check them."
	message_box("Error", mismatch_files +" these files are mismatched")
	add_error_log("These files are mismatched " + mismatch_files)

else: #if files are grouped, this statement starts follow-up table process.
	print "Follow-up table process starting."
	if not is_exist_source_xl: #EMS table not found.
		print source_xl + " named files not found."
		message_box("Error", source_xl + " named files not found.")
		add_error_log(source_xl + " named files not found.")
	
	elif not is_exist_destination_xl: #Follow-up table not found
		print destination_xl + " named files not found."
		message_box("Error", destination_xl + " named files not found.")
		add_error_log(destination_xl + " named files not found.")
	
	else: #if excel tables are found in folder, will proceed to next stages of process.
		remove_file_permission(destination_xl)
		print "File permission granted."

		indus_number = get_indus_number()
		print "Indus number defined."

		print "Data collecting belong to selected indus number."
		found_in_xl, delivery, ecu_type, sw_version, cr_number, cr_creation = get_from_xl(indus_number)
		print "Data collected belong to selected indus number."

		missing_value = check_missing_value(delivery, ecu_type, sw_version, cr_number, cr_creation)
		print "Checked missing values."

		print "Checking whether it was appended before."
		appended_before = check_appended_before(indus_number, sw_version)
		print "Checking completed."

		#---CHECKING THAT FOLDER CONSIST ONE S19 AND DOCX FILE---#
		folder_list = filter(lambda x: os.path.isdir(x), os.listdir('.')) #list of folders in main folder
		for folders in folder_list:
			correct_inside_of_folders = check_num_of_files(folders, 1)
			if not correct_inside_of_folders:
				break
		#---CHECKING THAT FOLDER CONSIST ONE S19 AND DOCX FILE---#	
		
		correct_sw_version = check_sw_version(sw_version)        


		if found_in_xl: #if the given indus number is found in EMS table.
			
			if not missing_value: #if the given indus number doesn't consist of missing values.
				
				if not appended_before: #if the given indus number with ther sw code is not appended before.
					
					if correct_inside_of_folders: #if folders consist only one s19 and docx file.
						
						if correct_sw_version: #if sw version matched with files in folders.
							print "Datas appending to Follow-up table."
							for folders in folder_list:
								
								file_list = os.listdir(folders)

								for file_names in file_list:
									if file_names.endswith(".s19") or file_names.endswith(".S19"):
										comment = file_names
									else:
										doc_name = folders + '/' + file_names
										title, ref_daimler, sw_title, sw_ref_daimler = reach_to_docx(doc_name, False)
								
								item_quantity += 1
								add_list(cr_number, cr_creation, sw_version, indus_number, ecu_type, title, ref_daimler, item_quantity, comment)

							#---SOFTWARE UPDATE JOB---#
							if delivery == 1: 
								add_list(cr_number, cr_creation, sw_version, indus_number, ecu_type, sw_title, sw_ref_daimler, 1, code_folder_name)
							#---SOFTWARE UPDATE JOB---#	
							
							to_excel() #

							#---CREATE CODE FOLDER---#
							if not os.path.exists("CODE") or not os.path.exists("Code"): #creates code folder if no code folder was created previously
								create_code_folder()
								print "Code folder was created and copied freigabe file into."
							#---CREATE CODE FOLDER---#	

						else:
							message_box("Error", "Please check Error-Logs file.")
							add_error_log(sw_version + " sotware version in excel table does'nt match with files.")
					else:    
						message_box("Error", "Please check Error-Logs file.")
						add_error_log('Problem founded in folder. Please check ' + folders + ' named folder.')
				else: 
					message_box("Error", "Please check Error-Logs file.")
					add_error_log("Indus " +indus_number+ " with " + sw_version + " software code appended before to excel")
			else:
				message_box("Error", "Please check Error-Logs file.") 
				add_error_log(source_xl + " file does not contain a record for Indus " + indus_number)              
		else:
			message_box("Error", "Please check Error-Logs file.")
			add_error_log("Indus number could'nt find in excel sheet.")

print "Process Completed"
message_box('Complete', 'Process Completed.')            
add_error_log("Time:" + str(datetime.datetime.now().time()))
