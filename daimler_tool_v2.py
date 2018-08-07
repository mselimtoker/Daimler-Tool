# coding=utf-8
from docx import Document
import shutil
import os
import subprocess
import stat
import datetime
import pyexcel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import ctypes  # An included library with Python install.

print "Modules imported."

destination_xl = 'Follow-up_CR_CN_ECU_Daimler_solenoide_piezo.xlsx'
destination_sheet_save_as = 'Follow-up_CR_CN_ECU_Daimler_solenoide_piezo.xlsx'
destination_sheet_name = 'DAIMLER 651'
code_folder_name = "CODE"

print "Constants defined."

item_quantity = 0
values_to_xl = []
blank_row = (' ', ' ', ' ', ' ', ' ')


def reach_to_docx(doc_name, grouping):

    if doc_name.endswith('.docx'):
        document = Document(doc_name)
        tables = document.tables

        if not tables[0].rows[31].cells[2].text:

            if not tables[0].rows[32].cells[2].text:

                if not tables[0].rows[33].cells[2].text:

                    if not tables[0].rows[34].cells[2].text:
                        key_position = 0
                        with open("Error-Logs.txt", "a") as text_file:
                            text_file.write(doc_name + " file has no code.\n")
                            text_file.write(
                                "-----------------------------------------------------------------------------------\n")
                    else:
                        unique_key = tables[0].rows[34].cells[2].text
                        key_position = 4
                else:
                    unique_key = tables[0].rows[33].cells[2].text
                    key_position = 3
            else:
                unique_key = tables[0].rows[32].cells[2].text
                key_position = 2

            if key_position != 0:
                with open("Error-Logs.txt", "a") as text_file:
                    text_file.write(doc_name + " 's code is in " + str(key_position) + ". line.\n")
                    text_file.write(
                        "------------------------------------------------------------------------------ \n")


        else:
            unique_key = tables[0].rows[31].cells[2].text
            key_position = 1

        if key_position != 0:
            unique_key_file_name = unique_key + '.s19'

            if not grouping:
                daimler_part_number = doc_name.split("/")[1].split("_")[1]
            else:
                daimler_part_number = doc_name.split("_")[1]
            print daimler_part_number

            file_name_first = "-".join(unique_key.split("-")[3:-1])
            file_name_sec = unique_key.split("-")[-1].split("_")[0]
            folder_name = file_name_first + '-' + file_name_sec + '-' + daimler_part_number

            sw_update_code = tables[0].rows[6].cells[3].text

            title = "-".join(unique_key.split("-")[0:3]) + '-' + daimler_part_number
            ref_daimler = daimler_part_number

            sw_ref_daimler = sw_update_code.replace(" ", "")
            sw_title = "-".join(unique_key.split("-")[0:3]) + '-' + sw_ref_daimler

    if grouping:
        return unique_key_file_name, folder_name, doc_name

    else:
        return title, ref_daimler, sw_title, sw_ref_daimler


def to_excel():
    destination_book = openpyxl.load_workbook(destination_xl)
    destination_sheet = destination_book[destination_sheet_name]
    print destination_xl + " found and opened."

    last_row = destination_sheet.max_row
    i = 1
    # <-----List Elements to Excel and Styling--------->

    for row in values_to_xl:
        destination_sheet.append(row)
        for col in range(1, 12):
            destination_sheet.cell(row=last_row + i, column=col).alignment = aligment_type
            destination_sheet.cell(row=last_row + i, column=col).border = my_border
            if delivery == 1:
                destination_sheet.cell(row=last_row + i, column=col).fill = delivery_color

        destination_sheet.cell(row=last_row + i, column=2).number_format = 'DD.MM.YYYY'

        destination_sheet.cell(row=last_row + i, column=21).alignment = aligment_type
        destination_sheet.cell(row=last_row + i, column=21).border = my_border
        if delivery == 1:
            destination_sheet.cell(row=last_row + i, column=21).fill = delivery_color

        i += 1

    destination_sheet.append(blank_row)

    for col in range(1, 12):
        destination_sheet.cell(row=last_row + i, column=col).fill = blank_row_color
        destination_sheet.cell(row=last_row + i, column=col).border = my_border

    destination_sheet.cell(row=last_row + i, column=21).fill = blank_row_color
    destination_sheet.cell(row=last_row + i, column=21).border = my_border

    # <----------------End of Blank Row Part----------->

    # <----------------Save File----------------------->

    destination_book.save(destination_sheet_save_as)


def create_folder_and_move(unique_key_file_name, folder_name, file_name):
    if os.path.exists(unique_key_file_name):
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

            shutil.move(unique_key_file_name, folder_name)
            shutil.move(file_name, folder_name)

    else:
        if not os.path.exists("Error-Logs.txt"):
            with open("Error-Logs.txt", "w") as text_file:
                text_file.write(file_name + " 's code dont match with\n    "
                                + unique_key_file_name + "\n")
                text_file.write(
                    "---------------------------------------------------------------------------\n")
        else:
            with open("Error-Logs.txt", "a") as text_file:
                text_file.write(file_name + " 's code dont match with\n    "
                                + unique_key_file_name + "\n")
                text_file.write(
                    "------------------------------------------------------------------------- \n")


def message_box(title, message):
    ctypes.windll.user32.MessageBoxA(None, message, title, 0)


def get_indus_number_from_folder(): #automatic version (read from folder names)
    indus_folder_name = os.getcwd().split("\\")[-1]

    if indus_folder_name.startswith("ME"):
        indus_number = indus_folder_name
    elif indus_folder_name.startswith("Indus"):
        indus_number = indus_folder_name.split("Indus")[-1]
    return int(indus_number) 


def get_indus_number(): #manuel version (user enter indus number manually)
	print "Please enter Indus Number: "
	i_num = raw_input()
	return i_num

def check_num_of_files(folder, num):
    num_of_docx = len([name for name in os.listdir(folder) if name.endswith('.docx')])
    num_of_s19 = len([name for name in os.listdir(folder) if name.endswith('.s19') or name.endswith('.S19')])

    if num_of_docx == num and num_of_s19 == num:
        return True
    else:
        return False


def check_appended_before(indus, sw_num):
    destination_book = openpyxl.load_workbook(destination_xl)
    destination_sheet = destination_book[destination_sheet_name]

    for i in range(1, destination_sheet.max_row):
        if destination_sheet.cell(row=i, column=5).value == sw_num and destination_sheet.cell(row=i,column=6).value == indus:
            append_before = True
            break
        else:
            append_before = False

    return append_before


def get_from_xl(indus):

    source_sheet = pyexcel.get_sheet(file_name='EMS_Meeting_Minutes_2018.xls')
    source_sheet.name_columns_by_row(0)
    print source_xl + " found and opened."

    for row in source_sheet:
        ## !!! Add same indus with diffrent sw code control part
        if row[3] == indus:
            found_in_xl_ = True
            ##delivery_ = row[2]
            ##ecu_type_ = row[4][3:]
            ##sw_version_ = row[6]
            ##cr_number_ = row[8]
            ##cr_creation_ = row[10]
            ##return True, row[2], row[4][3:], row[6], row[8], row[10]
            break

        else:
            found_in_xl_=False
    return found_in_xl_, row[2], row[4][3:], row[6], row[8], row[10]


def set_row_color(color_code):
    return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')


def set_border_style(style):
    return Border(left=Side(style=style), right=Side(style=style), top=Side(style=style), bottom=Side(style=style))


def set_aligment (horizontal, vertical):
    return openpyxl.styles.Alignment(horizontal=horizontal,vertical=horizontal)


def remove_file_permission(file):
    os.chmod(file, stat.S_IWRITE)
    print "File read and write permission granted."


def add_list(cr_num, cr_crea, sw_ver, indus, ecu, titl, ref_daim, item_quantity, comment):
    values_to_xl.append((cr_num, cr_crea, '', '', sw_ver, indus, ecu, titl, ref_daim, '', item_quantity, '', '', '', '', '', '', '', '', '', comment))


def check_missing_value(*args):
    for arg in args:
        if arg == "" or arg ==" ":
            result = True
            break
        else:
          result = False
    return result


def create_code_folder():
    src = filter(lambda x: os.path.isdir(x), os.listdir('.'))
    dest = "CODE"
    shutil.copytree(src[0], dest)


def check_sw_version(sw_ver):
    src = filter(lambda x: os.path.isdir(x), os.listdir('.'))
    f_name = src[0]
    for files in os.listdir(f_name):
        if files.endswith('.s19') or  files.endswith('.S19'):
            s19_sw = files.split("-")[2]
            if sw_ver == s19_sw:
                return True
            else:
                return False


def convert_doc_to_docx(path):
	ffff = os.listdir(path)
	for fi in ffff:
		if fi.endswith('.doc'):
			subprocess.call(['C:\Program Files\LibreOffice\program\soffice.exe', '--headless', '--convert-to', 'docx', fi])
        	os.remove(fi)

def find_ems_table_name():
	for file in os.listdir(os.curdir):
		if file.startswith("EMS_Meeting_Minutes_"):
			return file
			break


source_xl = find_ems_table_name()

my_border = set_border_style('thin') 
aligment_type = set_aligment('center', 'center') 

blank_row_color = set_row_color('a6a6a6') 
delivery_color = set_row_color('fde9d9') 

convert_doc_to_docx(os.curdir)

is_require_grouping = not check_num_of_files(os.curdir, 0) 
is_exist_excel_files = os.path.isfile(source_xl) and os.path.isfile(destination_xl) 

if is_require_grouping:
    print("Files are grouping")
    for file_list in os.listdir(os.curdir):
        if file_list.endswith('.docx'):
            s19, fold, docx = reach_to_docx(file_list, True)
            create_folder_and_move(s19, fold, docx)
    print("Files grouping process finished.")

is_ready_for_to_xl = check_num_of_files(os.curdir, 0)

if not is_ready_for_to_xl:
    print("There are non-grouped files. Please check them.")
    message_box("Error", "There are non-grouped files. Please check them.")

else:
    if not is_exist_excel_files:
        print "Excel files not found."
        message_box("Error", "Excel files not found.")

    else:
        remove_file_permission(destination_xl)
        print("File permission granted.")

        indus_number = get_indus_number()
        print("Indus number defined.")

        found_in_xl, delivery, ecu_type, sw_version, cr_number, cr_creation = get_from_xl(indus_number)
        print "Data collected belong to selected indus number."

        missing_value = check_missing_value(delivery, ecu_type, sw_version, cr_number, cr_creation)
        print "Checked missing values."

        appended_before = check_appended_before(indus_number, sw_version)
        print "Checked appended before."

        if found_in_xl:
            if missing_value:
                message_box("Error", "Selected Indus number has missing values.")
            elif appended_before:
                message_box("Error", "Indus: " +indus_number+ "with " + sw_version + " software code appended before to excel")
        else:
            message_box("Error", "Indus number could'nt find in excel sheet.")

        is_everything_correct = found_in_xl and not appended_before and not missing_value

        if not is_everything_correct:
            print("Error. Error. Error. Error. Error.")
        else:
            folder_list = filter(lambda x: os.path.isdir(x), os.listdir('.'))
            print "Folder paths listed."

            for folders in folder_list:
                correct_inside_of_folders = check_num_of_files(folders, 1)

                if correct_inside_of_folders:
                    file_list = os.listdir(folders)

                    for file_names in file_list:
                        if file_names.endswith(".s19") or file_names.endswith(".S19"):
                            comment = file_names
                        else:
                            doc_name = folders + '/' + file_names
                            title, ref_daimler, sw_title, sw_ref_daimler = reach_to_docx(doc_name, False)

                    if check_sw_version(sw_version):
                        item_quantity += 1
                        add_list(cr_number, cr_creation, sw_version, indus_number, ecu_type,
                                 title, ref_daimler, item_quantity, comment)

                    else:
                        print "Sw version not correct."

                else:
                    message_box("Error", 'Problem founded in folder. Please check ' + folders + ' named folder.')
            if delivery == 1: 
                add_list(cr_number, cr_creation, sw_version, indus_number, ecu_type,
                         sw_title, sw_ref_daimler, 1, code_folder_name)
                if not os.path.exists(code_folder_name):
                    create_code_folder()

            to_excel()
            message_box('Complete', 'Process Completed.')
with open("Error-Logs.txt", "a") as text_file:
    text_file.write("Time:" + str(datetime.datetime.now().time()) + "\n")
    text_file.write("-----------------------------------------------------------------------------------------------\n")
